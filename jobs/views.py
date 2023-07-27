from datetime import datetime
from django.shortcuts import render
from . models import Working
from django.db.models import Sum
import io
from django.http import FileResponse, HttpResponse
from reportlab.pdfgen import canvas
import openpyxl

def index(request):
    thour=0
    otime=0
    if request.method=='POST':
        start=request.POST['s_date']
        end=request.POST['e_date']
        work=Working.objects.filter(jdate__range=[start, end])
        for a in work:
            s1 = str(a.start)
            s2 = str(a.end)
            start_dt = datetime.strptime(s1, '%H:%M:%S')
            end_dt = datetime.strptime(s2, '%H:%M:%S')
            diff = (end_dt - start_dt)
            days = diff.days
            days_to_hours = days * 24
            diff_btw_two_times = (diff.seconds) / 3600
            overall_hours = days_to_hours + diff_btw_two_times
            thour=thour+overall_hours
            if a.lunch=='Yes' and a.holyday=='No' and overall_hours>7:
                overall_hours=overall_hours-1
                otime=overall_hours-7
            if a.lunch=='No' and a.holyday=='Yes':
                otime=overall_hours
            if a.lunch=='No' and a.holyday=='No':
                otime=overall_hours
            if a.lunch=='Yes' and a.holyday=='Yes':
                overall_hours=overall_hours-1
                otime=overall_hours
            print("WH",overall_hours)
        print("TT",thour)
        print("TOT",otime)
        context= {
            'work': work
            }
        return render(request, "index.html", context)
    return render(request, "index.html")
def some_view(request):
    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    # Create the PDF object, using the buffer as its "file."
    p = canvas.Canvas(buffer)

    # Draw things on the PDF. Here's where the PDF generation happens.
    # See the ReportLab documentation for the full list of functionality.
    p.drawString(100, 100, "Hello world.")

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="hello.pdf")
def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'My Data'

    # Write header row
    header = ['ID', 'Name', 'Email']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = Working.objects.all().values_list('name', 'jdate', 'start')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response